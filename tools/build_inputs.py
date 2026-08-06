#!/usr/bin/env python3
"""Turn the partner MASTER xlsx into generator inputs.

Outputs (into the project root):
  partners.csv           -> company,domain,contact_name,contact_email
  partners.enriched.json -> { "<lowercased company>": {firmographics...} }

Usage:
  python tools/build_inputs.py "<path-to.xlsx>"

Only firmographic fields are pulled (industry, size, location, revenue,
partner type, best contact). The migration-specific drip columns from the
prior Laplink campaign are intentionally ignored.
"""
import sys, os, json, csv
from openpyxl import load_workbook

def clean(v):
    if v is None: return ""
    s = str(v).strip()
    return "" if s.upper() in ("N/A", "NONE", "") else s

def pick(row, *names):
    for n in names:
        v = clean(row.get(n))
        if v: return v
    return ""

def main():
    if len(sys.argv) < 2:
        print("usage: python tools/build_inputs.py <xlsx>"); sys.exit(1)
    path = sys.argv[1]
    here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip() if h is not None else "" for h in rows[0]]

    partners, enriched, seen = [], {}, set()
    for r in rows[1:]:
        row = {hdr[i]: r[i] for i in range(len(hdr))}
        company = clean(row.get("Company Name"))
        if not company: continue
        domain = clean(row.get("Domain"))
        name = pick(row, "BEST Contact", "C1 Name", "Web/Curated Name")
        email = pick(row, "BEST Email", "C1 Verified Email (post-check)",
                     "C1 Email", "Web/Curated Email")
        if email and "@" not in email: email = ""

        key = company.lower()
        if key in seen:  # de-dupe repeated companies, first wins
            continue
        seen.add(key)

        partners.append({"company": company, "domain": domain,
                         "contact_name": name, "contact_email": email})

        city, state = clean(row.get("City")), clean(row.get("State"))
        location = ", ".join([p for p in (city, state) if p])
        rec = {}
        if location: rec["location"] = location
        for src, dst in (("Industry","industry"), ("Employees","size"),
                         ("Revenue","revenue"), ("Partner Type","partner_type"),
                         ("BEST Title","contact_title")):
            v = clean(row.get(src))
            if v: rec[dst] = v
        enriched[key] = rec

    with open(os.path.join(here, "partners.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["company","domain","contact_name","contact_email"])
        w.writeheader(); w.writerows(partners)
    with open(os.path.join(here, "partners.enriched.json"), "w", encoding="utf-8") as f:
        json.dump(enriched, f, indent=2, ensure_ascii=False)

    print(f"partners.csv: {len(partners)} rows")
    print(f"partners.enriched.json: {len(enriched)} companies")
    with_email = sum(1 for p in partners if p['contact_email'])
    print(f"rows with a contact email: {with_email}/{len(partners)}")

if __name__ == "__main__":
    main()
